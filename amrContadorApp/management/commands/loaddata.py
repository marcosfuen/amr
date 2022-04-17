import datetime
import email
import glob
import imaplib
import os
import zipfile
from urllib.parse import urlparse

import psycopg2
from decouple import config
from django.core.management.base import BaseCommand
from django.db import OperationalError
from openpyxl import load_workbook


class Command(BaseCommand):
    help = 'Se conecta el correo y descraga los ficheros del dia con las lectura de los ARM'

    def add_arguments(self, parser):
        # Named arguments
        parser.add_argument('--mail-url', default=config('MAIL_URL'))
        parser.add_argument('--db-url', default=config('DATABASE_URL'))
        parser.add_argument('--path-all-file-xlsx', default=config('PATH_All_FILE_XLSX'))

    def handle(self, *args, **options):
        mail_url = urlparse(options['mail_url'])
        mail_host = mail_url.hostname
        mail_host_port = mail_url.port
        mail_user = mail_url.username
        mail_user_pass = mail_url.password
        db_url = urlparse(options['db_url'])
        db_name = db_url.path.split('/', 1)[-1]
        db_host = db_url.hostname
        db_host_port = db_url.port
        db_user = db_url.username
        db_user_pass = db_url.password
        path_all_file_xlsx = options['path_all_file_xlsx']

        mail = imaplib.IMAP4_SSL(mail_host, mail_host_port)
        mail.login(mail_user, mail_user_pass)
        mail.select('Inbox')
        date = imapDateText(datetime.date.today())

        # este para la primera vez
        # type, data = mail.search(None, 'ALL')
        # este es para ponerlo cada 6 horas
        # type, data = mail.search(None, ('UNSEEN'), '(SENTSINCE {0})'.format(date))
        # type, data = mail.search(['SINCE', datetime.today()])
        # type, data = mail.uid('search', None, '(OR UNSEEN (SINCE %s))' % date)
        # solo los correos no leidos
        type, data = mail.search(None, '(UNSEEN)')
        mail_ids = data[0]
        id_list = mail_ids.split()

        for num in data[0].split():
            typ, data = mail.fetch(num, '(RFC822)')
            raw_email = data[0][1]  # converts byte literal to string removing b''
            raw_email_string = raw_email.decode('utf-8')
            email_message = email.message_from_string(
                raw_email_string)  # downloading attachments
            for part in email_message.walk():
                # this part comes from the snipped I don't understand yet...
                if part.get_content_maintype() == 'multipart':
                    continue
                if part.get('Content-Disposition') is None:
                    continue
                fileName = part.get_filename()
                if bool(fileName):
                    filePath = os.path.join(path_all_file_xlsx+'/', fileName)
                    if not os.path.isfile(filePath):
                        fp = open(filePath, 'wb')
                        fp.write(part.get_payload(decode=True))
                        fp.close()
                        subject = str(email_message).split(
                            "Subject: ", 1)[1].split("\nTo:", 1)[0]
        # termina la descarga de los ficheros del dia

        # se procesa los exel y se sube a la bd
        try:

            # ruta de la carpeta donde esta todos los fichero los compatados para descomprimir y los xlsx a renombrar
            path = path_all_file_xlsx
            # Extrae todos los ficheros .zip del directorio
            contenidoZip = os.listdir(path)
            # recorer todo los ficheros .zip para descomprimir
            zipFile = []
            password = None
            for ficheroZip in contenidoZip:
                if os.path.isfile(os.path.join(path, ficheroZip)) and ficheroZip.endswith('.zip'):
                    nombreFileZip = path+'/'+ficheroZip
                    archivo_zip = zipfile.ZipFile(nombreFileZip, "r")
                    try:
                        archivo_zip.extractall(pwd=password, path=path)
                    except:
                        pass
                    archivo_zip.close()
            # Renombrar todos los xlsx
            files = os.listdir(path)
            for index, file in enumerate(files):
                if os.path.isfile(os.path.join(path, file)) and file.endswith('.xlsx'):
                    os.rename(os.path.join(path, file), os.path.join(
                        path, ''.join([str(index), '.xlsx'])))
            # ruta completa del fichero
            # rt = os.path.join(os.path.dirname(__file__), config('NAME_FILE'))
            # conexion a la bd
            database = None
            try:
                database = psycopg2.connect(dbname=db_name, user=db_user, password=db_user_pass, host=db_host, port=db_host_port)
            except OperationalError as conx:
                self.stderr.write(self.style.ERROR("Cannot connect to db"))
                exit(1)
            # creamos el cursor
            cursor = database.cursor()
            # consulta para insertar los datos en la bd
            query = """ INSERT INTO "g_AMRContador" ("region", "numeroCuenta", "numeroCliente", "direccionDispositivo", "numeroContador", "nombreContador") VALUES (%s, %s, %s, %s, %s, %s) """
            queryDatos = """ INSERT INTO "g_Dato" ("amrContador_id", "tmpLectura", "totalImpEnerg", "t1ImpEnerg", "t2ImpEnerg", "t3ImpEnerg", "totalExpEnerg", "t1ExpEnerg", "t2ExpEnerg", "t3ExpEnerg", "totalImpReacEnerg", "t1ImpMaxDeman", "t2ImpMaxDeman", "t3ImpMaxDeman", "t1ExpMaxDeman", "t2ExpMaxDeman", "t3ExpMaxDeman", "imporActPoder", "exporActPoder", "factPoder", "faceAVolt", "faceBVolt", "faceCVolt", "faceACurr", "faceBCurr", "faceCCurr") VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s) """
            # listas de rutas de todos los xlsx a leer para importar
            # la ruta del directorios de los ficheros la tien el path asi que la utilizaremos
            # tomamos el contenido del directorio
            contenido = os.listdir(path)
            # lista con las direciones de mis xlsx
            xlsx = []
            # recorremos el contenido
            for fichero in contenido:
                if os.path.isfile(os.path.join(path, fichero)) and fichero.endswith('.xlsx'):
                    xlsx.append(path+'/'+fichero)
            # recorrer lista de xlsx para leer cada uno su contenido
            # lista de numero de contadores para que no esten duplicados
            numeroContadors = []
            for nombre in xlsx:
                # Leer fichero
                wb = load_workbook(nombre)
                # Seleccionar hoja activa
                ws = wb.active
                for index, row in enumerate(ws.iter_rows(min_row=2, min_col=0, max_col=32, max_row=ws.max_row)):
                    if row[4].value not in numeroContadors:
                        numeroContadors.append(row[4].value)
                        region = row[0].value
                        numeroCuenta = row[1].value
                        numeroCliente = row[2].value
                        direccionDispositivo = row[3].value
                        numeroContador = row[4].value
                        nombreContador = row[5].value
                        values = (region, numeroCuenta, numeroCliente,
                                  direccionDispositivo, numeroContador, nombreContador)
                        try:
                            cursor.execute(query, values)
                            database.commit()
                        except BaseException as exc:
                            self.stderr.write(self.style.ERROR("se jodio esto import AMR"))
                            database.commit()
                        # self.stdout.write((row[0].value, row[1].value, row[2].value,
                        #       row[3].value, row[4].value, row[5].value)
            cursor.execute(""" SELECT * FROM "g_AMRContador" """)
            rows = cursor.fetchall()
            for nombre in xlsx:
                # Leer fichero
                wb = load_workbook(nombre)
                # Seleccionar hoja activa
                ws = wb.active
                for row in rows:
                    for index, row1 in enumerate(ws.iter_rows(min_row=2, min_col=0, max_col=32, max_row=ws.max_row)):
                        if row[5] == row1[4].value:
                            amrContador_id = row[0]
                            tmpLectura = row1[6].value
                            totalImpEnerg = row1[8].value
                            t1ImpEnerg = row1[9].value
                            t2ImpEnerg = row1[10].value
                            t3ImpEnerg = row1[11].value
                            totalExpEnerg = row1[12].value
                            t1ExpEnerg = row1[13].value
                            t2ExpEnerg = row1[14].value
                            t3ExpEnerg = row1[15].value
                            totalImpReacEnerg = row1[16].value
                            t1ImpMaxDeman = row1[17].value
                            t2ImpMaxDeman = row1[18].value
                            t3ImpMaxDeman = row1[19].value
                            t1ExpMaxDeman = row1[20].value
                            t2ExpMaxDeman = row1[21].value
                            t3ExpMaxDeman = row1[22].value
                            imporActPoder = row1[23].value
                            exporActPoder = row1[24].value
                            factPoder = row1[25].value
                            faceAVolt = row1[26].value
                            faceBVolt = row1[27].value
                            faceCVolt = row1[28].value
                            faceACurr = row1[29].value
                            faceBCurr = row1[30].value
                            faceCCurr = row1[31].value
                            values = (amrContador_id, tmpLectura, totalImpEnerg,
                                      t1ImpEnerg, t2ImpEnerg, t3ImpEnerg, totalExpEnerg,
                                      t1ExpEnerg, t2ExpEnerg, t3ExpEnerg, totalImpReacEnerg,
                                      t1ImpMaxDeman, t2ImpMaxDeman, t3ImpMaxDeman, t1ExpMaxDeman,
                                      t2ExpMaxDeman, t3ExpMaxDeman, imporActPoder, exporActPoder,
                                      factPoder, faceAVolt, faceBVolt, faceCVolt, faceACurr, faceBCurr, faceCCurr)

                            cursor.execute(queryDatos, values)
                            database.commit()

            # self.stdout.write(numeroContador)
            # cerramos el curso
            cursor.close()
            # hacemos en commit para persistir los datos en la bd
            # database.commit()
            # cerramos la conexion con la bd
            database.close()
            # borramos todos los ficheros
            filesRemove = glob.glob(path+'/*')
            for f in filesRemove:
                os.remove(f)
        except Exception as exc:
            # OpenPyXL no puede parsear el fichero
            self.stderr.write(self.style.ERROR(exc))
            exit(1)


def imapDateText(dt):
    months = ('Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
              'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec')
    return '%02d-%s-%04d' % (dt.day, months[dt.month - 1], dt.year)
